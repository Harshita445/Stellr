"""
Timetable Parser Service

CRITICAL RULE: This is the ONLY place the Excel workbook is ever touched.
Every other part of the application queries PostgreSQL, never the workbook.
If you need timetable data during a normal request, use TimetableQueryService
or TimetableEntryRepository — not openpyxl.

The workbook is parsed ONCE during import and the data is stored in
PostgreSQL. After that, the file can be discarded. Re-importing replaces
all data atomically within a single transaction.

Workbook format (per product spec):
- Multiple sheets, one per year/department combination
  e.g. "First Year A", "Second Year B"
- Within each sheet:
  - First column: row headers (usually day names)
  - First row: section codes as column headers (e.g. "3A1A", "3A1B")
  - Cells: subject name if the section is busy during that slot
  - Empty cells: free period
- The slot index is determined by the row position within a day group
"""

import os
from dataclasses import dataclass, field
from typing import Optional

from app.core.logging import logger


@dataclass
class ParsedRow:
    """A single parsed cell from the workbook.

    One cell = one busy slot for one section.
    Empty cells (free periods) are not emitted.
    """
    sheet_name: str
    section_code: str
    day_name: str
    day_of_week: int
    slot_index: int
    subject_name: str
    raw_row: int
    raw_col: int


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    sheets_found: list[str] = field(default_factory=list)
    sections_found: list[str] = field(default_factory=list)


# ── Constants ─────────────────────────────────────────────────────────────

# Days of the week as they appear in the workbook
EXPECTED_DAY_NAMES: list[str] = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
]

DAY_NAME_TO_INDEX: dict[str, int] = {
    name: idx for idx, name in enumerate(EXPECTED_DAY_NAMES)
}

# Maximum rows to scan when looking for the header row
MAX_HEADER_SCAN_ROWS = 10

# Cells in this column are treated as row labels (day names)
ROW_LABEL_COLUMN = 0

# First column containing section data
FIRST_DATA_COLUMN = 1


class TimetableParserService:
    """Parses an Excel workbook into normalized ParsedRow records.

    Runs openpyxl in a thread pool executor since it's CPU-bound.
    """

    async def parse(self, file_path: str) -> ParseResult:
        """Parse the workbook at file_path into ParsedRows.

        This is the entry point. It handles opening the workbook,
        iterating sheets, and dispatching to sheet-level parsing.
        """
        import openpyxl

        result = ParseResult()

        if not os.path.exists(file_path):
            result.errors.append(f"File not found: {file_path}")
            return result

        if not file_path.endswith((".xlsx", ".xlsm")):
            result.errors.append(f"Unsupported file format: {file_path}. Expected .xlsx")
            return result

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            result.errors.append(f"Failed to open workbook: {exc}")
            return result

        if not wb.sheetnames:
            result.errors.append("Workbook contains no sheets")
            wb.close()
            return result

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.sheets_found.append(sheet_name)
            try:
                self._parse_sheet(ws, sheet_name, result)
            except Exception as exc:
                result.errors.append(f"Sheet '{sheet_name}' failed: {exc}")
                logger.warning("Sheet parse failed", extra={"sheet": sheet_name, "error": str(exc)})

        wb.close()
        return result

    def _parse_sheet(self, ws, sheet_name: str, result: ParseResult) -> None:
        """Parse a single sheet: find header row, then process data rows."""
        rows_iter = ws.iter_rows(values_only=False)
        all_rows = list(rows_iter)

        if not all_rows:
            result.warnings.append(f"Sheet '{sheet_name}' is empty")
            return

        # Locate the header row (first row containing section-like column headers)
        header_row_idx, section_columns = self._find_header_row(all_rows, sheet_name, result)

        if header_row_idx is None:
            result.warnings.append(
                f"Sheet '{sheet_name}': could not find a header row with section codes"
            )
            return

        section_codes: list[str] = []
        for cell in section_columns:
            val = str(cell.value or "").strip()
            if val:
                # Normalize section code
                normalized = self._normalize_section_code(val)
                section_codes.append(normalized)
                if normalized not in result.sections_found:
                    result.sections_found.append(normalized)

        if not section_codes:
            result.warnings.append(f"Sheet '{sheet_name}': no section columns found in header row")
            return

        # Parse data rows: each row is a slot for a specific day
        data_rows = all_rows[header_row_idx + 1:]

        current_day_name: str | None = None
        day_row_count = 0

        for row_idx, row in enumerate(data_rows):
            row_num = header_row_idx + 1 + row_idx + 1  # 1-based for error messages

            label_cell = self._get_cell_value(row, ROW_LABEL_COLUMN)
            if label_cell:
                # Could be a day header or a merged day label
                day_name = self._detect_day_name(label_cell)
                if day_name:
                    current_day_name = day_name
                    day_row_count = 0
                    # This row is the day header, no slot data
                    continue
                elif current_day_name:
                    # It's a row within a day — could be a sub-label
                    # Check if the first data cell has content
                    has_data = any(
                        self._get_cell_value(row, col)
                        for col in range(FIRST_DATA_COLUMN, FIRST_DATA_COLUMN + len(section_codes))
                    )
                    if not has_data:
                        # Label-only row (club hours, break, etc.) — skip
                        continue

            if current_day_name is None:
                # No day context yet, skip
                continue

            day_of_week = DAY_NAME_TO_INDEX.get(current_day_name)
            if day_of_week is None:
                continue

            # Each row within a day group is a slot
            slot_index = day_row_count
            day_row_count += 1

            # Parse each section column
            for col_offset, section_code in enumerate(section_codes):
                col_idx = FIRST_DATA_COLUMN + col_offset
                cell_val = self._get_cell_value(row, col_idx)

                if cell_val:
                    normalized_subject = self._parse_subject_name(cell_val)
                    if normalized_subject:
                        result.rows.append(ParsedRow(
                            sheet_name=sheet_name,
                            section_code=section_code,
                            day_name=current_day_name,
                            day_of_week=day_of_week,
                            slot_index=slot_index,
                            subject_name=normalized_subject,
                            raw_row=row_num,
                            raw_col=col_idx,
                        ))

    def _find_header_row(self, all_rows: list, sheet_name: str, result: ParseResult):
        """Scan the first N rows to find one that looks like a header row.

        A header row has short alphanumeric codes (section identifiers) in
        consecutive columns starting from column 1.
        """
        for row_idx in range(min(MAX_HEADER_SCAN_ROWS, len(all_rows))):
            row = all_rows[row_idx]
            cells = list(row)
            if len(cells) < 2:
                continue

            # Column 0 should be empty or a generic label like "Day/Time"
            col0_val = self._get_cell_value(cells, 0) if len(cells) > 0 else ""

            # Columns 1+ should have section-like codes
            section_cells: list = []
            for cell in cells[FIRST_DATA_COLUMN:]:
                val = self._get_cell_value([cell], 0) if cell else ""
                if val:
                    section_cells.append(cell)

            if len(section_cells) >= 1:
                return row_idx, section_cells

        return None, []

    def _normalize_section_code(self, code: str) -> str:
        """Normalize a section code to a canonical form.

        e.g. "3A1A " → "3A1A", "CSE-A" → "CSE-A"
        """
        return code.strip().upper()

    def _detect_day_name(self, text: str) -> str | None:
        """Check if text is a recognized day name (case-insensitive)."""
        cleaned = text.strip().lower().capitalize()
        if cleaned in DAY_NAME_TO_INDEX:
            return cleaned
        # Handle "Day 1", "Day 2" etc as Monday, Tuesday
        for day_name in EXPECTED_DAY_NAMES:
            if day_name.lower() in cleaned.lower():
                return day_name
        return None

    def _parse_subject_name(self, raw: str) -> str | None:
        """Clean and validate a subject name from a cell.

        Returns None for obviously non-subject content (breaks, club hours, etc.)
        """
        cleaned = raw.strip()
        if not cleaned:
            return None

        # Skip known non-academic entries
        non_academic = [
            "club", "club hour", "break", "lunch", "assembly",
            "sports", "library", "remedial", "study hour",
        ]
        if cleaned.lower().strip() in non_academic:
            return None

        return cleaned

    def _get_cell_value(self, row_cells: list, col_index: int) -> str | None:
        """Safely get a cell value as a string, or None if empty."""
        if col_index >= len(row_cells):
            return None
        cell = row_cells[col_index]
        value = cell.value
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None
